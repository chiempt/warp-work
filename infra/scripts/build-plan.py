#!/usr/bin/env python3
"""Render docs/planning/backlog.csv into a formatted spreadsheet.

The CSV is the source of truth: it is what pull requests touch, what review
comments can point at, and what git can merge. The spreadsheet is a *view* —
convenient to filter and read, useless to review. Never edit it directly; the
next `make plan` would discard the change.

The workbook deliberately holds only Overview and Backlog. Open questions and
architecture decisions already have a home in docs/open-questions.md and
docs/decisions/, and copying them here would be a second place to keep in step.
"""
import csv
import pathlib
import sys
from collections import Counter
from datetime import date

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except ModuleNotFoundError:
    sys.exit(
        "openpyxl is not installed, and it is only needed to rebuild the\n"
        "spreadsheet view. The backlog itself is docs/planning/backlog.csv,\n"
        "which opens in any editor or spreadsheet.\n\n"
        "  pip3 install openpyxl"
    )

CSV_PATH = pathlib.Path("docs/planning/backlog.csv")
OUT_PATH = pathlib.Path("docs/planning/warp-delivery-plan.xlsx")

DONE, WIP, TODO, BLOCKED, DECIDE = (
    "Done", "In Progress", "Not Started", "Blocked", "Decision Needed",
)
STATUSES = [DONE, WIP, TODO, BLOCKED, DECIDE]
PRIORITIES = ["P0", "P1", "P2", "P3"]

FILL = {DONE: "C6EFCE", WIP: "FFEB9C", TODO: "F2F2F2", BLOCKED: "FFC7CE", DECIDE: "DDEBF7"}
INK = {DONE: "1E6B34", WIP: "8A6100", TODO: "595959", BLOCKED: "9C2531", DECIDE: "1F4E79"}
MEANING = {
    DONE: "Built and verified in the repository.",
    WIP: "Partly built; the row says which part.",
    TODO: "Declared or planned, no implementation.",
    BLOCKED: "Waiting on an open question, not on effort.",
    DECIDE: "Waiting on a decision from the maintainer.",
}
PRIORITY_MEANING = {
    "P0": "Blocks the phase. Nothing above it works without this.",
    "P1": "Needed for the phase to be usable.",
    "P2": "Improves the phase; can follow it.",
    "P3": "Deferred deliberately.",
}

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def read_rows():
    if not CSV_PATH.exists():
        sys.exit(f"{CSV_PATH} not found")
    with CSV_PATH.open(newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader)
        rows = [r for r in reader if any(r)]

    problems = []
    seen = set()
    for i, r in enumerate(rows, start=2):
        if len(r) != len(headers):
            problems.append(f"line {i}: {len(r)} fields, expected {len(headers)}")
            continue
        if r[0] in seen:
            problems.append(f"line {i}: duplicate ID {r[0]}")
        seen.add(r[0])
        if r[5] not in STATUSES:
            problems.append(f"line {i} ({r[0]}): unknown status {r[5]!r}")
        if r[6] not in PRIORITIES:
            problems.append(f"line {i} ({r[0]}): unknown priority {r[6]!r}")

    # Dependencies must exist, or the plan quietly points at nothing.
    for i, r in enumerate(rows, start=2):
        for dep in (d.strip() for d in r[7].split(",") if d.strip()):
            if dep.startswith("OQ-"):
                continue  # open questions live in docs/open-questions.md
            if dep not in seen:
                problems.append(f"line {i} ({r[0]}): depends on unknown ID {dep}")

    # A cycle makes the plan unusable: neither row can ever start, and reading
    # the file does not reveal it. Found one in practice - X-01 and C-01 each
    # waiting on the other.
    graph = {
        r[0]: [d.strip() for d in r[7].split(",") if d.strip() and not d.strip().startswith("OQ-")]
        for r in rows
    }
    WHITE, GREY, BLACK = 0, 1, 2
    colour = dict.fromkeys(graph, WHITE)

    def walk(node, path):
        colour[node] = GREY
        for dep in graph.get(node, ()):
            if colour.get(dep) == GREY:
                problems.append("dependency cycle: " + " -> ".join(path + [dep]))
            elif colour.get(dep) == WHITE:
                walk(dep, path + [dep])
        colour[node] = BLACK

    for node in graph:
        if colour[node] == WHITE:
            walk(node, [node])

    if problems:
        sys.exit("backlog.csv is not valid:\n  " + "\n  ".join(problems))

    return headers, rows


def head(ws, row, ncols, height=26):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = height


def build(headers, rows):
    wb = Workbook()

    ov = wb.active
    ov.title = "Overview"
    ov.sheet_view.showGridLines = False

    ov["B2"] = "Warp - Delivery Plan"
    ov["B2"].font = Font(bold=True, size=14, color="1F3864")
    ov["B3"] = "A personal work system: sees everything, tracks what is owed, drafts the work."
    ov["B3"].font = Font(size=10, color="595959")
    ov["B4"] = (
        f"Generated {date.today().isoformat()} from docs/planning/backlog.csv. "
        "Do not edit this file - edit the CSV and run `make plan`."
    )
    ov["B4"].font = Font(size=9, italic=True, color="808080")

    ov["B6"] = "Where things live"
    ov["B6"].font = Font(bold=True, size=11, color="1F3864")
    for i, (what, where) in enumerate([
        ("Backlog (source of truth)", "docs/planning/backlog.csv"),
        ("Open questions", "docs/open-questions.md"),
        ("Architecture decisions", "docs/decisions/"),
        ("What Warp is", "docs/warp-project-context.md"),
        ("How to contribute", "CONTRIBUTING.md"),
    ]):
        ov.cell(row=7 + i, column=2, value=what).font = Font(size=10)
        ov.cell(row=7 + i, column=3, value=where).font = Font(size=10, color="1F4E79")

    r = 14
    ov.cell(row=r, column=2, value="Status").font = Font(bold=True, size=11, color="1F3864")
    for i, s in enumerate(STATUSES):
        cell = ov.cell(row=r + 1 + i, column=2, value=s)
        cell.fill = PatternFill("solid", fgColor=FILL[s])
        cell.font = Font(bold=True, size=10, color=INK[s])
        cell.border = BORDER
        ov.cell(row=r + 1 + i, column=3, value=MEANING[s]).font = Font(size=10)

    r += 7
    ov.cell(row=r, column=2, value="Priority").font = Font(bold=True, size=11, color="1F3864")
    for i, p in enumerate(PRIORITIES):
        ov.cell(row=r + 1 + i, column=2, value=p).font = Font(bold=True, size=10)
        ov.cell(row=r + 1 + i, column=3, value=PRIORITY_MEANING[p]).font = Font(size=10)

    r += 6
    ov.cell(row=r, column=2, value="Progress by phase").font = Font(bold=True, size=11, color="1F3864")
    cols = ["Phase"] + STATUSES + ["Total", "% Done"]
    for i, h in enumerate(cols):
        ov.cell(row=r + 1, column=2 + i, value=h)
    head(ov, r + 1, 1 + len(cols))

    phases = list(dict.fromkeys(x[1] for x in rows))
    line = r + 2
    for phase in phases + ["ALL"]:
        subset = rows if phase == "ALL" else [x for x in rows if x[1] == phase]
        counts = Counter(x[5] for x in subset)
        bold = phase == "ALL"
        ov.cell(row=line, column=2, value=phase).font = Font(size=10, bold=bold)
        for i, s in enumerate(STATUSES):
            c = ov.cell(row=line, column=3 + i, value=counts.get(s, 0))
            c.alignment = Alignment(horizontal="center")
            c.font = Font(size=10, bold=bold)
        c = ov.cell(row=line, column=3 + len(STATUSES), value=len(subset))
        c.alignment = Alignment(horizontal="center")
        c.font = Font(size=10, bold=bold)
        pc = ov.cell(row=line, column=4 + len(STATUSES),
                     value=counts.get(DONE, 0) / len(subset) if subset else 0)
        pc.number_format = "0%"
        pc.alignment = Alignment(horizontal="center")
        pc.font = Font(size=10, bold=bold)
        for cc in range(2, 5 + len(STATUSES)):
            ov.cell(row=line, column=cc).border = BORDER
        line += 1

    for col, w in {"A": 3, "B": 30, "C": 60, "D": 12, "E": 12, "F": 12,
                   "G": 15, "H": 9, "I": 9}.items():
        ov.column_dimensions[col].width = w

    bl = wb.create_sheet("Backlog")
    bl.append(headers)
    head(bl, 1, len(headers))
    for row in rows:
        bl.append(row)

    for i in range(2, bl.max_row + 1):
        status = bl.cell(row=i, column=6).value
        for c in range(1, len(headers) + 1):
            cell = bl.cell(row=i, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 5))
            cell.font = Font(size=10)
        sc = bl.cell(row=i, column=6)
        sc.fill = PatternFill("solid", fgColor=FILL[status])
        sc.font = Font(size=10, bold=True, color=INK[status])
        sc.alignment = Alignment(vertical="top", horizontal="center")
        bl.cell(row=i, column=1).font = Font(size=10, bold=True)
        bl.cell(row=i, column=7).alignment = Alignment(vertical="top", horizontal="center")
        bl.row_dimensions[i].height = 46

    for col, w in {"A": 8, "B": 15, "C": 14, "D": 34, "E": 78,
                   "F": 15, "G": 9, "H": 16, "I": 34}.items():
        bl.column_dimensions[col].width = w

    bl.freeze_panes = "B2"
    bl.auto_filter.ref = f"A1:{chr(ord('A') + len(headers) - 1)}{bl.max_row}"

    dv = DataValidation(type="list", formula1=f'"{",".join(STATUSES)}"', allow_blank=False)
    dv.errorTitle, dv.error = "Unknown status", "Use one of the five statuses from the Overview sheet."
    bl.add_data_validation(dv)
    dv.add(f"F2:F{bl.max_row}")

    dvp = DataValidation(type="list", formula1=f'"{",".join(PRIORITIES)}"', allow_blank=False)
    bl.add_data_validation(dvp)
    dvp.add(f"G2:G{bl.max_row}")

    wb.save(OUT_PATH)
    return Counter(x[5] for x in rows)


if __name__ == "__main__":
    headers, rows = read_rows()
    counts = build(headers, rows)
    done = counts.get(DONE, 0)
    print(f"{OUT_PATH}: {len(rows)} rows, {done} done ({done / len(rows) * 100:.0f}%)")
